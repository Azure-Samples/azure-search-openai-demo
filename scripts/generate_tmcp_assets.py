# pyright: reportGeneralTypeIssues=false

"""Generate TMCP v2.4 supporting assets (Excel workbooks and PNG diagrams).

Run this script after installing `openpyxl` and `Pillow` in the active environment.
"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parents[1]
ASSET_DIR = ROOT / "docs" / "tmcp"
ASSET_DIR.mkdir(parents=True, exist_ok=True)

PRIMARY_FONT = "DejaVuSans.ttf"


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
	try:
		return ImageFont.truetype(PRIMARY_FONT, size)
	except OSError:
		try:
			return ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", size)
		except OSError:
			return ImageFont.load_default()


def stylize_header(cells: Iterable[Cell]) -> None:
	for cell in cells:
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center")
		cell.fill = PatternFill(start_color="F0F6FF", end_color="F0F6FF", fill_type="solid")


def auto_width(ws: Worksheet, mapping: dict[str, float]) -> None:
	for column, width in mapping.items():
		ws.column_dimensions[column].width = width


def build_scenario_calculator(path: Path) -> None:
	wb = Workbook()
	ws = cast(Worksheet, wb.active)
	ws.title = "Calculator"
	ws["A1"] = "Metric"
	ws["B1"] = "Value"
	stylize_header(ws["A1":"B1"][0])

	entries: list[tuple[str, float | str]] = [
		("Process (P)", 0.84),
		("Architecture (A)", 0.88),
		("Quality Slack (Qs)", 0.20),
		("Q", "=1+B4"),
		("Non-Blocker Score", 6.0),
		("CAP", 20),
		("κ (Blocker Impact)", 1),
		("Blockers", 0),
		("Base", "=B2*B3"),
		("R*", "="),
	]

	cell_map: dict[str, str] = {}
	for idx, (label, value) in enumerate(entries, start=2):
		cell = f"B{idx}"
		cell_map[label] = cell
		ws[f"A{idx}"] = label
		ws[cell] = value

	ws[cell_map["Q"]] = f"=1+{cell_map['Quality Slack (Qs)']}"
	ws[cell_map["Base"]] = f"={cell_map['Process (P)']}*{cell_map['Architecture (A)']}"
	ws[cell_map["R*"]] = (
		f"=({cell_map['Base']}^{cell_map['Q']})"
		f"-({cell_map['Non-Blocker Score']}/{cell_map['CAP']})"
		f"-({cell_map['κ (Blocker Impact)']}*{cell_map['Blockers']})"
	)

	ws["D1"] = "Gate"
	ws["E1"] = "Threshold"
	ws["F1"] = "Value"
	ws["G1"] = "Status"
	stylize_header(ws["D1":"G1"][0])
	gates: list[tuple[str, str, float, str]] = [
		("Base ≥ 0.75", "Base", 0.75, ">="),
		("Q ≥ 1.20", "Q", 1.20, ">="),
		("R* ≥ 0.10", "R*", 0.10, ">="),
		("Blockers", "Blockers", 0, "<="),
	]
	for idx, (label, ref, threshold, op) in enumerate(gates, start=2):
		ws[f"D{idx}"] = label
		ws[f"E{idx}"] = threshold
		ws[f"F{idx}"] = f"={cell_map[ref]}"
		if op == ">=":
			ws[f"G{idx}"] = f"=IF(F{idx}>=E{idx},\"PASS\",\"FAIL\")"
		else:
			ws[f"G{idx}"] = f"=IF(F{idx}<=E{idx},\"PASS\",\"FAIL\")"

	auto_width(ws, {"A": 26, "B": 20, "D": 20, "E": 14, "F": 16, "G": 14})

	scenario_ws = cast(Worksheet, wb.create_sheet("Scenarios"))
	headers = [
		"Scenario",
		"Process (P)",
		"Architecture (A)",
		"Qs",
		"Q",
		"NBS",
		"Blockers",
		"R*",
	]
	scenario_ws.append(headers)
	stylize_header(scenario_ws["A1":"H1"][0])
	scenarios: list[tuple[str, float, float, float, float, int]] = [
		("Legacy Ops", 0.58, 0.55, 0.05, 12, 1),
		("Stabilizing", 0.72, 0.70, 0.12, 9, 0),
		("Balanced", 0.83, 0.86, 0.18, 6, 0),
		("A11y Push", 0.79, 0.90, 0.22, 5, 0),
		("Hardening", 0.90, 0.94, 0.28, 3, 0),
	]
	last_row = 1
	for idx, (name, p_val, a_val, qs_val, nbs_val, blockers_val) in enumerate(scenarios, start=2):
		scenario_ws[f"A{idx}"] = name
		scenario_ws[f"B{idx}"] = p_val
		scenario_ws[f"C{idx}"] = a_val
		scenario_ws[f"D{idx}"] = qs_val
		scenario_ws[f"E{idx}"] = f"=1+D{idx}"
		scenario_ws[f"F{idx}"] = nbs_val
		scenario_ws[f"G{idx}"] = blockers_val
		scenario_ws[f"H{idx}"] = (
			f"=((B{idx}*C{idx})^E{idx})-(F{idx}/{cell_map['CAP']})-(G{idx}*{cell_map['κ (Blocker Impact)']})"
		)
		last_row = idx

	chart = LineChart()
	chart.title = "Scenario R* Trend"
	chart.style = 3
	data = Reference(scenario_ws, min_col=8, min_row=1, max_row=last_row)
	chart.add_data(data, titles_from_data=True)
	cats = Reference(scenario_ws, min_col=1, min_row=2, max_row=last_row)
	chart.set_categories(cats)
	chart.y_axis.title = "R*"
	chart.x_axis.title = "Scenario"
	scenario_ws.add_chart(chart, "J2")
	auto_width(scenario_ws, {"A": 20, "B": 14, "C": 14, "D": 10, "E": 10, "F": 10, "G": 10, "H": 14})

	wb.save(path)


def build_scorecard(path: Path) -> None:
	wb = Workbook()
	ws = cast(Worksheet, wb.active)
	ws.title = "Scorecard"
	ws.merge_cells("A1:C1")
	ws["A1"] = "TMCP Scorecard v2.2"
	ws["A1"].font = Font(bold=True, size=16)
	ws["A1"].alignment = Alignment(horizontal="left")

	process_rows: list[tuple[str, float, float]] = [
		("Branch Coverage", 0.34, 0.85),
		("Definition of Done", 0.33, 0.82),
		("PR Labels", 0.33, 0.78),
	]
	ws["A2"] = "Process Signals"
	ws["A2"].font = Font(bold=True)
	ws["A3"] = "Metric"
	ws["B3"] = "Weight"
	ws["C3"] = "Value"
	stylize_header(ws["A3":"C3"][0])
	row = 4
	for metric, weight, value in process_rows:
		ws[f"A{row}"] = metric
		ws[f"B{row}"] = weight
		ws[f"C{row}"] = value
		row += 1
	p_row = row
	ws[f"A{p_row}"] = "Process (P)"
	ws[f"C{p_row}"] = f"=SUMPRODUCT(B4:B{p_row-1},C4:C{p_row-1})"
	p_cell = f"C{p_row}"

	row = p_row + 2
	ws[f"A{row-1}"] = "Architecture Signals"
	ws[f"A{row-1}"].font = Font(bold=True)
	ws[f"A{row}"] = "Metric"
	ws[f"B{row}"] = "Weight"
	ws[f"C{row}"] = "Value"
	stylize_header(ws[f"A{row}:C{row}"][0])
	row += 1
	architecture_rows: list[tuple[str, float, float]] = [
		("Security Scans", 0.4, 0.88),
		("Low-Risk Ratio", 0.2, 0.74),
		("Tests + Coverage", 0.4, 0.86),
	]
	arch_start = row
	for metric, weight, value in architecture_rows:
		ws[f"A{row}"] = metric
		ws[f"B{row}"] = weight
		ws[f"C{row}"] = value
		row += 1
	a_row = row
	ws[f"A{a_row}"] = "Architecture (A)"
	ws[f"C{a_row}"] = f"=SUMPRODUCT(B{arch_start}:B{a_row-1},C{arch_start}:C{a_row-1})"
	a_cell = f"C{a_row}"

	qs_row = a_row + 2
	ws[f"A{qs_row}"] = "Quality Slack (Qs)"
	ws[f"C{qs_row}"] = 0.20
	qs_cell = f"C{qs_row}"
	nbs_row = qs_row + 1
	ws[f"A{nbs_row}"] = "Non-Blocker Score (NBS)"
	ws[f"C{nbs_row}"] = 6
	nbs_cell = f"C{nbs_row}"
	cap_row = nbs_row + 1
	ws[f"A{cap_row}"] = "CAP"
	ws[f"C{cap_row}"] = 20
	cap_cell = f"C{cap_row}"
	kappa_row = cap_row + 1
	ws[f"A{kappa_row}"] = "κ (Blocker Impact)"
	ws[f"C{kappa_row}"] = 1
	kappa_cell = f"C{kappa_row}"
	blockers_row = kappa_row + 1
	ws[f"A{blockers_row}"] = "Blockers"
	ws[f"C{blockers_row}"] = 0
	blockers_cell = f"C{blockers_row}"
	base_row = blockers_row + 1
	ws[f"A{base_row}"] = "Base (P×A)"
	ws[f"C{base_row}"] = f"={p_cell}*{a_cell}"
	base_cell = f"C{base_row}"
	q_row = base_row + 1
	ws[f"A{q_row}"] = "Q"
	ws[f"C{q_row}"] = f"=1+{qs_cell}"
	q_cell = f"C{q_row}"
	r_row = q_row + 1
	ws[f"A{r_row}"] = "R*"
	ws[f"C{r_row}"] = f"=({base_cell}^{q_cell})-({nbs_cell}/{cap_cell})-({kappa_cell}*{blockers_cell})"
	r_cell = f"C{r_row}"

	ws["E3"] = "Gate"
	ws["F3"] = "Value"
	ws["G3"] = "Threshold"
	ws["H3"] = "Status"
	stylize_header(ws["E3":"H3"][0])
	gate_rows: list[tuple[str, str, float, str]] = [
		("Base ≥ 0.75", base_cell, 0.75, ">="),
		("Q ≥ 1.20", q_cell, 1.20, ">="),
		("R* ≥ 0.10", r_cell, 0.10, ">="),
		("Blockers", blockers_cell, 0, "<="),
	]
	gate_start = 4
	for idx, (label, ref, threshold, op) in enumerate(gate_rows, start=gate_start):
		ws[f"E{idx}"] = label
		ws[f"F{idx}"] = f"={ref}"
		ws[f"G{idx}"] = threshold
		if op == ">=":
			ws[f"H{idx}"] = f"=IF(F{idx}>=G{idx},\"PASS\",\"FAIL\")"
		else:
			ws[f"H{idx}"] = f"=IF(F{idx}<=G{idx},\"PASS\",\"FAIL\")"
	verdict_row = gate_start + len(gate_rows)
	ws[f"E{verdict_row}"] = "Verdict"
	ws[f"F{verdict_row}"] = "=IF(COUNTIF(H4:H7,\"FAIL\")=0,\"APPROVE\",\"REJECT\")"

	auto_width(ws, {"A": 28, "B": 12, "C": 18, "E": 22, "F": 18, "G": 14, "H": 12})

	reference_ws = cast(Worksheet, wb.create_sheet("Reference"))
	reference_ws["A1"] = "Severity Weights"
	stylize_header(reference_ws["A1":"B1"][0])
	severities = [
		("critical", 5.0),
		("high", 3.0),
		("medium", 1.5),
		("low", 1.0),
		("info", 0.5),
	]
	for idx, (name, weight) in enumerate(severities, start=2):
		reference_ws[f"A{idx}"] = name
		reference_ws[f"B{idx}"] = weight
	reference_ws["D2"] = "Formula"
	reference_ws["D3"] = "R* = (P × A)^Q − (NBS / CAP) − κ · Blockers"
	reference_ws["D5"] = "Usage"
	reference_ws["D6"] = "Use the Scenarios sheet to compare what-if inputs."
	reference_ws["D7"] = "Link"
	reference_ws["D8"] = "docs/tmcp/TMCP_Scenario_Calculator.xlsx"
	auto_width(reference_ws, {"A": 16, "B": 12, "D": 48})

	wb.save(path)


def _line_height(font: ImageFont.ImageFont) -> int:
	bbox = font.getbbox("Hy")
	return bbox[3] - bbox[1]


def create_formula_image(path: Path) -> None:
	width, height = 1400, 520
	image = Image.new("RGB", (width, height), "#0f172a")
	draw = ImageDraw.Draw(image)
	title_font = load_font(52)
	formula_font = load_font(44)
	text_font = load_font(28)
	draw.text((60, 40), "TMCP v2.4 Risk Formula", font=title_font, fill="#ffffff")
	draw.text((60, 130), "R* = (P × A)^Q − (NBS / CAP) − κ · Blockers", font=formula_font, fill="#34d399")
	bullets = [
		"P: Process signal from branches, Definition of Done, labels",
		"A: Architecture signal from scans, low-risk ratio, tests + coverage",
		"Q = 1 + Qs where Qs blends coverage, tests, risk slack",
		"NBS: Weighted score of non-blocking findings, bounded by CAP",
		"κ: Blocker coefficient (1). Any blocker subtracts directly from R*.",
	]
	y = 210
	spacing = _line_height(text_font) + 10
	for bullet in bullets:
		draw.text((80, y), f"• {bullet}", font=text_font, fill="#e2e8f0")
		y += spacing
	image.save(path)


def create_steps_image(path: Path) -> None:
	width, height = 1400, 640
	image = Image.new("RGB", (width, height), "#eef2ff")
	draw = ImageDraw.Draw(image)
	title_font = load_font(48)
	step_font = load_font(32)
	body_font = load_font(26)
	draw.text((50, 30), "TMCP v2.4 Delivery Steps", font=title_font, fill="#1e3a8a")
	steps = [
		("Aggregate", "Collect SARIF/JSON outputs (Bandit, pip-audit, detect-secrets)."),
		("Normalize", "Run tmcp_aggregate.py to build summary.json + weights."),
		("Compute Signals", "Feed summary + coverage into tmcp_compute_pa.py."),
		("Evaluate Gates", "Apply Base/Q/R*/Blocker thresholds with emoji summary."),
		("Plan Scenarios", "Tune the Scenario Calculator & Scorecard workbooks."),
	]
	y = 120
	for idx, (title, body) in enumerate(steps, start=1):
		draw.text((70, y), f"{idx}. {title}", font=step_font, fill="#1d4ed8")
		y += _line_height(step_font) + 6
		draw.text((110, y), body, font=body_font, fill="#1f2937")
		y += _line_height(body_font) + 20
	draw.text(
		(70, height - 80),
		"Artifacts: tmcp_workflow_v2_4.yml · TMCP_Scenario_Calculator.xlsx · TMCP_Scorecard_v2.2.xlsx",
		font=body_font,
		fill="#4338ca",
	)
	image.save(path)


def main() -> None:
	build_scenario_calculator(ASSET_DIR / "TMCP_Scenario_Calculator.xlsx")
	build_scorecard(ASSET_DIR / "TMCP_Scorecard_v2.2.xlsx")
	create_formula_image(ASSET_DIR / "tmcp_formula.png")
	create_steps_image(ASSET_DIR / "tmcp_steps.png")
	print(f"Generated TMCP assets in {ASSET_DIR}")


if __name__ == "__main__":
	main()
